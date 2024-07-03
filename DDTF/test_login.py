import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook  # Library for Excel file handling
from DDTF.login_page import LoginPage  # Assuming DDTF is the parent directory


@pytest.fixture
def setup():
    driver = webdriver.Chrome()
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    yield driver
    driver.quit()


def test_login_ddt(setup, excel_data):
    username = excel_data["Username"]
    password = excel_data["Password"]
    login_page = LoginPage(setup)

    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login()

    if login_page.is_login_successful():
        test_result = "Pass"
    else:
        test_result = "Fail"

    # Update Excel sheet (replace with your specific sheet and cell references)
    wb = load_workbook("login_data.xlsx")  # Assuming Excel file is in the same directory
    sheet = wb["Sheet1"]  # Replace with your sheet name
    sheet.cell(row=excel_data["Test ID"] + 1, column=7).value = test_result
    wb.save("login_data.xlsx")


@pytest.mark.parametrize("excel_data", read_excel_data("login_data.xlsx"))  # Define fixture for data
def read_excel_data(file):
    wb = load_workbook(file)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2):
        data.append({"Test ID": row[0].value, "Username": row[1].value, "Password": row[2].value})
    return data
